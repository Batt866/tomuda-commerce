from __future__ import annotations

import io
import json
import tempfile
from datetime import date, timedelta
from pathlib import Path

from django.test import TestCase, override_settings

from dashboard.api import _retained_orders_state
from dashboard.models import AppState, ProductImage
from dashboard.product_images import (
    find_stored_product_image_url,
    hydrate_product_images,
    persist_imported_product_images,
    save_product_image_bytes,
)
from dashboard.profile_images import (
    find_stored_profile_image_url,
    hydrate_profile_images,
    save_profile_image_bytes,
)
from dashboard.seed_data import default_state


@override_settings(SECURE_SSL_REDIRECT=False)
class ProductImageStorageTests(TestCase):
    def test_product_thumb_is_generated_and_served(self):
        from PIL import Image

        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=Path(tmp), MEDIA_URL="/media/"):
                buf = io.BytesIO()
                Image.new("RGB", (640, 480), (30, 120, 200)).save(buf, format="JPEG", quality=90)
                raw = buf.getvalue()
                url = save_product_image_bytes("thumb-prod", raw, "jpg")

                self.assertIn("/media/products/thumb-prod.jpg", url)
                thumb_file = Path(tmp) / "products" / "thumb-prod_t.jpg"
                self.assertTrue(thumb_file.is_file())
                self.assertLess(thumb_file.stat().st_size, len(raw))

                response = self.client.get("/media/products/thumb-prod_t.jpg")
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response["Content-Type"], "image/jpeg")
                body = (
                    b"".join(response.streaming_content)
                    if hasattr(response, "streaming_content")
                    else response.content
                )
                self.assertGreater(len(body), 32)

                # On-demand rebuild when thumb file is missing.
                thumb_file.unlink()
                rebuilt = self.client.get("/media/products/thumb-prod_t.jpg")
                self.assertEqual(rebuilt.status_code, 200)
                rebuilt_body = (
                    b"".join(rebuilt.streaming_content)
                    if hasattr(rebuilt, "streaming_content")
                    else rebuilt.content
                )
                self.assertGreater(len(rebuilt_body), 32)
                self.assertTrue(thumb_file.is_file())

    def test_oversized_jpeg_is_compressed_instead_of_rejected(self):
        import os

        from PIL import Image

        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=Path(tmp), MEDIA_URL="/media/"):
                buf = io.BytesIO()
                Image.frombytes(
                    "RGB", (1200, 1200), os.urandom(1200 * 1200 * 3)
                ).save(buf, format="JPEG", quality=95)
                raw = buf.getvalue()
                self.assertGreater(len(raw), 300_000)
                url = save_product_image_bytes("big-prod", raw, "jpg")
                self.assertIn("/media/products/big-prod.jpg", url)
                stored = ProductImage.objects.get(product_id="big-prod")
                self.assertLessEqual(len(stored.image), 300_000)
                self.assertGreater(len(stored.image), 32)

    def test_serves_png_when_jpg_url_is_requested(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=Path(tmp), MEDIA_URL="/media/"):
                raw = b"png-bytes-for-ext-mismatch" * 3
                save_product_image_bytes("ext-prod", raw, "png")
                response = self.client.get("/media/products/ext-prod.jpg")
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response["Content-Type"], "image/png")
                body = (
                    b"".join(response.streaming_content)
                    if hasattr(response, "streaming_content")
                    else response.content
                )
                self.assertEqual(body, raw)

    def test_product_image_serves_from_db_when_media_file_is_missing(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=Path(tmp), MEDIA_URL="/media/"):
                raw = b"image-bytes-for-db-fallback" * 3
                url = save_product_image_bytes("prod-1", raw, "png")

                self.assertIn("/media/products/prod-1.png", url)
                self.assertTrue(ProductImage.objects.filter(product_id="prod-1").exists())

                media_file = Path(tmp) / "products" / "prod-1.png"
                media_file.unlink()

                stored_url = find_stored_product_image_url("prod-1")
                self.assertIn("/media/products/prod-1.png", stored_url)

                response = self.client.get("/media/products/prod-1.png")
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response["Content-Type"], "image/png")
                self.assertEqual(response.content, raw)

    def test_hydrate_keeps_media_url_when_storage_missing(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=Path(tmp), MEDIA_URL="/media/"):
                state = {
                    "products": [
                        {
                            "id": "missing-prod",
                            "name": "Missing",
                            "image": "/media/products/missing-prod.jpg?v=1",
                        }
                    ]
                }

                next_state, changed = hydrate_product_images(state)

                self.assertFalse(changed)
                self.assertEqual(
                    next_state["products"][0]["image"],
                    "/media/products/missing-prod.jpg?v=1",
                )

    def test_existing_media_file_is_backfilled_to_db(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=Path(tmp), MEDIA_URL="/media/"):
                media_dir = Path(tmp) / "products"
                media_dir.mkdir(parents=True)
                raw = b"existing-media-file" * 3
                (media_dir / "old-prod.jpg").write_bytes(raw)

                url = find_stored_product_image_url("old-prod")

                self.assertIn("/media/products/old-prod.jpg", url)
                image = ProductImage.objects.get(product_id="old-prod")
                self.assertEqual(bytes(image.image), raw)
                self.assertEqual(image.content_type, "image/jpeg")

    def test_persist_with_empty_product_ids_leaves_other_images_untouched(self):
        # Importing rows that only update existing products (e.g. a price-only
        # Excel) produces an empty product_ids list. That must NOT reprocess the
        # whole catalog and clear images whose media file is missing on prod.
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=Path(tmp), MEDIA_URL="/media/"):
                state = {
                    "products": [
                        {
                            "id": "untouched-prod",
                            "name": "Untouched",
                            "image": "/media/products/untouched-prod.jpg?v=1",
                        }
                    ]
                }

                report = persist_imported_product_images(
                    state,
                    previous_state=state,
                    product_ids=[],
                )

                self.assertEqual(report["imageSkipped"], 0)
                self.assertEqual(
                    state["products"][0]["image"],
                    "/media/products/untouched-prod.jpg?v=1",
                )

    def test_employee_image_serves_from_db_when_media_file_is_missing(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=Path(tmp), MEDIA_URL="/media/"):
                raw = b"employee-photo-bytes" * 3
                url = save_profile_image_bytes("employee", "emp-1", raw, "jpg")

                self.assertIn("/media/employees/emp-1.jpg", url)
                media_file = Path(tmp) / "employees" / "emp-1.jpg"
                media_file.unlink()

                stored_url = find_stored_profile_image_url("employee", "emp-1")
                self.assertIn("/media/employees/emp-1.jpg", stored_url)

                response = self.client.get("/media/employees/emp-1.jpg")
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.content, raw)

    def test_hydrate_profile_keeps_media_url_when_storage_missing(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=Path(tmp), MEDIA_URL="/media/"):
                state = {
                    "employees": [
                        {
                            "id": "emp-1",
                            "name": "Test",
                            "image": "/media/employees/emp-1.jpg?v=1",
                        }
                    ]
                }

                next_state, changed = hydrate_profile_images(state)

                self.assertFalse(changed)
                self.assertEqual(
                    next_state["employees"][0]["image"],
                    "/media/employees/emp-1.jpg?v=1",
                )

    def test_save_state_returns_cleaned_state(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=Path(tmp), MEDIA_URL="/media/"):
                product = {
                    "id": "missing-prod",
                    "barcode": "123",
                    "name": "Missing",
                    "category": "Бусад",
                    "unit": "ширхэг",
                    "boxQuantity": 1,
                    "price": 1000,
                    "costPrice": 0,
                    "stock": 0,
                    "minStock": 0,
                    "country": "Монгол",
                    "image": "",
                }
                initial = default_state()
                initial["products"] = [dict(product)]
                state = default_state()
                state["products"] = [
                    {**product, "image": "/media/products/missing-prod.jpg?v=1"}
                ]
                AppState.objects.create(key="main", data=initial)

                response = self.client.post(
                    "/api/state",
                    {
                        "state": state,
                        "actor": {"id": "admin", "email": "admin@tomuda.mn"},
                    },
                    content_type="application/json",
                )

                self.assertEqual(response.status_code, 200)
                payload = response.json()
                self.assertEqual(
                    payload["state"]["products"][0]["image"],
                    "/media/products/missing-prod.jpg?v=1",
                )


@override_settings(SECURE_SSL_REDIRECT=False)
class OrderRetentionTests(TestCase):
    def test_retained_orders_state_drops_orders_older_than_thirty_days(self):
        today = date.today()
        state = default_state()
        state["settings"]["orderRetentionDays"] = 30
        state["orders"] = [
            {
                "id": "old",
                "createdAt": (today - timedelta(days=31)).isoformat(),
                "status": "delivered",
                "paymentTerm": "cash",
                "isPaid": True,
            },
            {
                "id": "recent",
                "createdAt": (today - timedelta(days=10)).isoformat(),
                "status": "delivered",
                "paymentTerm": "cash",
                "isPaid": True,
            },
        ]

        cleaned = _retained_orders_state(state)

        self.assertEqual([o["id"] for o in cleaned["orders"]], ["recent"])

    def test_retained_orders_keeps_unpaid_receivable_past_retention(self):
        today = date.today()
        state = default_state()
        state["settings"]["orderRetentionDays"] = 30
        state["orders"] = [
            {
                "id": "old-unpaid",
                "createdAt": (today - timedelta(days=90)).isoformat(),
                "status": "delivered",
                "paymentTerm": "credit",
                "isPaid": False,
            },
            {
                "id": "old-paid",
                "createdAt": (today - timedelta(days=90)).isoformat(),
                "status": "delivered",
                "paymentTerm": "credit",
                "isPaid": True,
            },
        ]

        cleaned = _retained_orders_state(state)

        self.assertEqual([o["id"] for o in cleaned["orders"]], ["old-unpaid"])

    def test_get_state_purges_expired_orders_from_database(self):
        today = date.today()
        state = default_state()
        state["settings"]["orderRetentionDays"] = 30
        state["orders"] = [
            {
                "id": "old",
                "createdAt": (today - timedelta(days=40)).isoformat(),
                "status": "delivered",
                "paymentTerm": "cash",
                "isPaid": True,
            },
            {
                "id": "recent",
                "createdAt": (today - timedelta(days=5)).isoformat(),
                "status": "delivered",
                "paymentTerm": "cash",
                "isPaid": True,
            },
        ]
        AppState.objects.create(key="main", data=state)

        response = self.client.get("/api/state")

        self.assertEqual(response.status_code, 200)
        order_ids = [o["id"] for o in response.json()["state"]["orders"]]
        self.assertEqual(order_ids, ["recent"])

        row = AppState.objects.get(key="main")
        saved_ids = [o["id"] for o in row.data["orders"]]
        self.assertEqual(saved_ids, ["recent"])

    def test_health_purges_expired_orders_from_database(self):
        today = date.today()
        state = default_state()
        state["settings"]["orderRetentionDays"] = 30
        state["orders"] = [
            {
                "id": "old",
                "createdAt": (today - timedelta(days=45)).isoformat(),
                "status": "delivered",
                "paymentTerm": "cash",
                "isPaid": True,
            }
        ]
        AppState.objects.create(key="main", data=state)

        response = self.client.get("/api/health")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["purgedOrders"], 1)
        row = AppState.objects.get(key="main")
        self.assertEqual(row.data["orders"], [])


class ProductExcelImportTests(TestCase):
    def test_template_headers_exclude_cost_and_image(self):
        from dashboard.excel_import import PRODUCT_HEADERS, build_product_template_bytes

        self.assertNotIn("Үйлдвэрлэсэн үнэ", PRODUCT_HEADERS)
        self.assertNotIn("Зураг URL", PRODUCT_HEADERS)
        content = build_product_template_bytes()
        self.assertTrue(content.startswith(b"PK"))

    def test_import_skips_template_example_and_creates_new_rows(self):
        from dashboard.excel_import import import_products_into_state

        state = default_state()
        state["products"] = []
        rows = [
            ["Barcode", "Барааны нэр", "Хэмжих нэгж", "Борлуулалтын үнэ", "Үйлдвэрлэсэн улс"],
            ["6977236071316", "Жишээ бараа", "ширхэг", "15000", "Монгол"],
            ["1111111111111", "Шинэ бараа", "ширхэг", "12000", "Монгол"],
        ]
        next_state, report = import_products_into_state(state, rows)
        self.assertEqual(report["success"], 1)
        self.assertEqual(report["created"], 1)
        self.assertEqual(len(next_state["products"]), 1)
        self.assertEqual(next_state["products"][0]["name"], "Шинэ бараа")

    def test_import_accepts_edited_template_row(self):
        from dashboard.excel_import import import_products_into_state

        state = default_state()
        state["products"] = []
        rows = [
            ["Barcode", "Барааны нэр", "Хэмжих нэгж", "Борлуулалтын үнэ", "Үйлдвэрлэсэн улс"],
            ["5555555555555", "Миний бараа", "ширхэг", "18000", "Монгол"],
        ]
        next_state, report = import_products_into_state(state, rows)
        self.assertEqual(report["success"], 1)
        self.assertEqual(next_state["products"][0]["name"], "Миний бараа")

    def test_load_sheet_rows_sniffs_xlsx_without_extension(self):
        from dashboard.excel_import import build_product_template_bytes, load_sheet_rows

        content = build_product_template_bytes()
        rows = load_sheet_rows(content, "upload")
        self.assertTrue(rows)
        self.assertIn("Barcode", rows[0])

    def test_import_updates_existing_barcode_with_create_permission(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        from dashboard.excel_import import import_products_into_state
        from dashboard.permissions import validate_state_mutation

        state = default_state()
        state["products"] = [
            {
                "id": "prod-1",
                "barcode": "8888888888888",
                "name": "Хуучин",
                "category": "Бусад",
                "unit": "ширхэг",
                "price": 100,
                "costPrice": 0,
                "stock": 0,
                "minStock": 0,
                "country": "Монгол",
                "boxQuantity": 1,
                "image": "",
            }
        ]
        state["employees"] = [
            {
                "id": "emp-1",
                "email": "sales@test.mn",
                "name": "Sales",
                "role": "sales",
                "password": "x",
                "permissions": ["products.create"],
            }
        ]
        AppState.objects.update_or_create(key="main", defaults={"data": state})

        rows = [
            ["Barcode", "Барааны нэр", "Хэмжих нэгж", "Борлуулалтын үнэ", "Үйлдвэрлэсэн улс"],
            ["8888888888888", "Шинэчлэгдсэн", "ширхэг", "20000", "Монгол"],
            ["9999999999999", "Шинээр", "ширхэг", "30000", "Монгол"],
        ]
        current = AppState.objects.get(key="main").data
        next_state, report = import_products_into_state(current, rows)
        actor = {"id": "emp-1", "email": "sales@test.mn"}
        ok, message = validate_state_mutation(
            current, next_state, actor, import_kind="products"
        )
        self.assertTrue(ok, message)
        self.assertEqual(report["created"], 1)
        self.assertEqual(report["updated"], 1)

        actor_json = json.dumps(actor)
        buf = io.BytesIO()
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save(buf)
        upload = SimpleUploadedFile(
            "products.xlsx",
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(
            "/api/import/products",
            {"file": upload, "actor": actor_json},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report"]["success"], 2)
        names = {p["name"] for p in payload["products"]}
        self.assertIn("Шинэчлэгдсэн", names)
        self.assertIn("Шинээр", names)


class CustomerExcelImportTests(TestCase):
    def test_template_includes_latitude_longitude(self):
        from dashboard.excel_import import CUSTOMER_HEADERS, build_customer_template_bytes

        self.assertIn("Уртраг", CUSTOMER_HEADERS)
        self.assertIn("Өргөрөг", CUSTOMER_HEADERS)
        content = build_customer_template_bytes()
        self.assertTrue(content.startswith(b"PK"))

    def test_import_reads_customer_coordinates(self):
        from dashboard.excel_import import import_customers_into_state

        state = default_state()
        state["customers"] = []
        rows = [
            [
                "Нэр",
                "Регистр",
                "Утас 1",
                "Утас 2",
                "Аймаг / Хот",
                "Дүүрэг",
                "Хороо",
                "Дэлгэрэнгүй хаяг",
                "Уртраг",
                "Өргөрөг",
            ],
            [
                "Жишээ ХХК",
                "1234567",
                "99112233",
                "",
                "Улаанбаатар",
                "Баянзүрх",
                "1",
                "1-р хороо",
                "47.916935",
                "106.951798",
            ],
            [
                "GPS ХХК",
                "7654321",
                "88112233",
                "",
                "Улаанбаатар",
                "Сүхбаатар",
                "2",
                "2-р хороо",
                "47.925309",
                "106.930147",
            ],
        ]
        next_state, report = import_customers_into_state(state, rows)
        self.assertEqual(report["success"], 1)
        self.assertEqual(len(next_state["customers"]), 1)
        customer = next_state["customers"][0]
        self.assertEqual(customer["name"], "GPS ХХК")
        self.assertEqual(customer["latitude"], "47.925309")
        self.assertEqual(customer["longitude"], "106.930147")

    def test_customer_import_api_creates_rows(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from openpyxl import Workbook

        state = default_state()
        state["customers"] = []
        state["employees"] = [
            {
                "id": "emp-1",
                "email": "admin@test.mn",
                "name": "Admin",
                "role": "admin",
                "password": "x",
            }
        ]
        AppState.objects.update_or_create(key="main", defaults={"data": state})

        rows = [
            [
                "Нэр",
                "Регистр",
                "Утас 1",
                "Утас 2",
                "Аймаг / Хот",
                "Дүүрэг",
                "Хороо",
                "Дэлгэрэнгүй хаяг",
                "Уртраг",
                "Өргөрөг",
            ],
            [
                "GPS ХХК",
                "7654321",
                "88112233",
                "",
                "Улаанбаатар",
                "Сүхбаатар",
                "2",
                "2-р хороо",
                "47.925309",
                "106.930147",
            ],
        ]
        buf = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save(buf)
        upload = SimpleUploadedFile(
            "customers.xlsx",
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        actor = json.dumps({"id": "emp-1", "email": "admin@test.mn"})
        response = self.client.post(
            "/api/import/customers",
            {"file": upload, "actor": actor},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report"]["success"], 1)
        self.assertEqual(len(payload["customers"]), 1)
        self.assertEqual(payload["customers"][0]["name"], "GPS ХХК")


class StockInPermissionTests(TestCase):
    def test_warehouse_role_can_save_stock_in_without_products_edit(self):
        from dashboard.permissions import validate_state_mutation

        state = default_state()
        state["products"] = [
            {
                "id": "prod-1",
                "barcode": "111",
                "name": "Test",
                "category": "Бусад",
                "unit": "ширхэг",
                "price": 1000,
                "costPrice": 0,
                "stock": 10,
                "minStock": 0,
                "country": "Монгол",
                "boxQuantity": 1,
                "image": "",
            }
        ]
        state["employees"] = [
            {
                "id": "wh-1",
                "email": "wh@test.mn",
                "name": "Warehouse",
                "role": "warehouse",
                "password": "x",
            }
        ]
        state["stockInReceipts"] = []
        state["inventoryLogs"] = []

        next_state = json.loads(json.dumps(state))
        next_state["stockInReceipts"] = [
            {
                "id": "sir-1",
                "receiptNumber": "OR-202607-001",
                "receiptSeq": 1,
                "monthKey": "202607",
                "createdAt": "2026-07-09T00:00:00",
                "employeeId": "wh-1",
                "employeeName": "Warehouse",
                "lines": [
                    {
                        "productId": "prod-1",
                        "productName": "Test",
                        "quantity": 5,
                        "costPrice": 800,
                        "packs": 0,
                    }
                ],
                "totalAmount": 4000,
            }
        ]
        next_state["products"][0]["stock"] = 15
        next_state["products"][0]["costPrice"] = 800
        next_state["inventoryLogs"] = [
            {
                "id": "log-1",
                "productId": "prod-1",
                "productName": "Test",
                "type": "in",
                "quantity": 5,
                "costPrice": 800,
                "packs": 0,
                "date": "2026-07-09T00:00:00",
                "employeeName": "Warehouse",
                "receiptId": "sir-1",
                "receiptNumber": "OR-202607-001",
            }
        ]

        actor = {"id": "wh-1", "email": "wh@test.mn"}
        ok, message = validate_state_mutation(state, next_state, actor)
        self.assertTrue(ok, message)

    def test_stock_in_allows_unit_and_min_stock_normalization(self):
        from dashboard.permissions import validate_state_mutation

        state = default_state()
        state["products"] = [
            {
                "id": "prod-1",
                "barcode": "111",
                "name": "Test",
                "category": "Бусад",
                "unit": "KG",
                "price": 1000,
                "costPrice": 0,
                "stock": 10,
                "minStock": 0,
                "country": "Монгол",
                "boxQuantity": 1,
                "image": "",
            }
        ]
        state["employees"] = [
            {
                "id": "wh-1",
                "email": "wh@test.mn",
                "name": "Warehouse",
                "role": "warehouse",
                "password": "x",
            }
        ]
        state["stockInReceipts"] = []
        state["inventoryLogs"] = []

        next_state = json.loads(json.dumps(state))
        next_state["stockInReceipts"] = [
            {
                "id": "sir-1",
                "receiptNumber": "OR-202607-001",
                "receiptSeq": 1,
                "monthKey": "202607",
                "createdAt": "2026-07-09T00:00:00",
                "employeeId": "wh-1",
                "employeeName": "Warehouse",
                "lines": [
                    {
                        "productId": "prod-1",
                        "productName": "Test",
                        "quantity": 5,
                        "costPrice": 800,
                        "packs": 0,
                    }
                ],
                "totalAmount": 4000,
            }
        ]
        next_state["products"][0]["stock"] = 15
        next_state["products"][0]["costPrice"] = 800
        next_state["products"][0]["unit"] = "кг"
        next_state["products"][0]["minStock"] = 3
        next_state["inventoryLogs"] = [
            {
                "id": "log-1",
                "productId": "prod-1",
                "productName": "Test",
                "type": "in",
                "quantity": 5,
                "costPrice": 800,
                "packs": 0,
                "date": "2026-07-09T00:00:00",
                "employeeName": "Warehouse",
                "receiptId": "sir-1",
                "receiptNumber": "OR-202607-001",
            }
        ]

        actor = {"id": "wh-1", "email": "wh@test.mn"}
        ok, message = validate_state_mutation(state, next_state, actor)
        self.assertTrue(ok, message)

    def test_stock_in_api_persists_for_warehouse_role(self):
        state = default_state()
        state["products"] = [
            {
                "id": "prod-1",
                "barcode": "111",
                "name": "Test",
                "category": "Бусад",
                "unit": "ширхэг",
                "price": 1000,
                "costPrice": 0,
                "stock": 10,
                "minStock": 0,
                "country": "Монгол",
                "boxQuantity": 1,
                "image": "",
            }
        ]
        state["employees"] = [
            {
                "id": "wh-1",
                "email": "wh@test.mn",
                "name": "Warehouse",
                "role": "warehouse",
                "password": "x",
            }
        ]
        state["stockInReceipts"] = []
        state["inventoryLogs"] = []
        AppState.objects.update_or_create(key="main", defaults={"data": state})

        next_state = json.loads(json.dumps(state))
        next_state["stockInReceipts"] = [
            {
                "id": "sir-1",
                "receiptNumber": "OR-202607-001",
                "receiptSeq": 1,
                "monthKey": "202607",
                "createdAt": "2026-07-09T00:00:00",
                "employeeId": "wh-1",
                "employeeName": "Warehouse",
                "lines": [
                    {
                        "productId": "prod-1",
                        "productName": "Test",
                        "quantity": 5,
                        "costPrice": 800,
                        "packs": 0,
                    }
                ],
                "totalAmount": 4000,
            }
        ]
        next_state["products"][0]["stock"] = 15
        next_state["products"][0]["costPrice"] = 800
        next_state["inventoryLogs"] = [
            {
                "id": "log-1",
                "productId": "prod-1",
                "productName": "Test",
                "type": "in",
                "quantity": 5,
                "costPrice": 800,
                "packs": 0,
                "date": "2026-07-09T00:00:00",
                "employeeName": "Warehouse",
                "receiptId": "sir-1",
                "receiptNumber": "OR-202607-001",
            }
        ]

        response = self.client.post(
            "/api/state",
            {
                "state": next_state,
                "actor": {"id": "wh-1", "email": "wh@test.mn"},
            },
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200, response.content)
        saved = AppState.objects.get(key="main").data
        self.assertEqual(saved["products"][0]["stock"], 15)
        self.assertEqual(len(saved["stockInReceipts"]), 1)


@override_settings(SECURE_SSL_REDIRECT=False)
class MultiDeviceStateMergeTests(TestCase):
    def test_merge_keeps_customer_added_on_other_device(self):
        from dashboard.state_merge import merge_app_states

        remote = {
            "customers": [
                {"id": "c-a", "name": "Device A customer", "phone1": "99110000"},
                {"id": "c-shared", "name": "Shared", "phone1": "99000000"},
            ],
            "products": [],
            "employees": [],
            "orders": [],
            "deletionLog": [],
        }
        local = {
            "customers": [
                {"id": "c-shared", "name": "Shared updated", "phone1": "99000001"},
                {"id": "c-b", "name": "Device B customer", "phone1": "99220000"},
            ],
            "products": [],
            "employees": [],
            "orders": [],
            "deletionLog": [],
        }

        merged = merge_app_states(remote, local)
        ids = {c["id"] for c in merged["customers"]}
        self.assertEqual(ids, {"c-a", "c-b", "c-shared"})
        shared = next(c for c in merged["customers"] if c["id"] == "c-shared")
        self.assertEqual(shared["name"], "Shared updated")
        self.assertEqual(shared["phone1"], "99000001")

    def test_merge_keeps_newer_customer_fields_from_upsert(self):
        from dashboard.state_merge import merge_app_states

        remote = {
            "customers": [
                {
                    "id": "c-ulz",
                    "name": "Улз",
                    "latitude": 47.918,
                    "longitude": 106.917,
                    "updatedAt": "2026-08-08T10:00:00Z",
                }
            ],
            "products": [],
            "employees": [],
            "orders": [],
            "deletionLog": [],
        }
        # Stale peer full-state blob without the rename/map stamp.
        local = {
            "customers": [
                {
                    "id": "c-ulz",
                    "name": "Old store",
                    "latitude": "",
                    "longitude": "",
                }
            ],
            "products": [],
            "employees": [],
            "orders": [],
            "deletionLog": [],
        }

        merged = merge_app_states(remote, local)
        customer = next(c for c in merged["customers"] if c["id"] == "c-ulz")
        self.assertEqual(customer["name"], "Улз")
        self.assertEqual(customer["latitude"], 47.918)
        self.assertEqual(customer["longitude"], 106.917)
        self.assertEqual(customer["updatedAt"], "2026-08-08T10:00:00Z")

    def test_save_state_does_not_wipe_peer_device_customer(self):
        state = default_state()
        state["customers"] = [
            {
                "id": "c-device-a",
                "name": "From device A",
                "companyName": "A LLC",
                "phone1": "99111111",
                "registrationNumber": "",
                "address": "",
            }
        ]
        AppState.objects.create(key="main", data=state)

        # Device B posts an older snapshot that never saw c-device-a,
        # but adds its own customer.
        device_b = default_state()
        device_b["customers"] = [
            {
                "id": "c-device-b",
                "name": "From device B",
                "companyName": "B LLC",
                "phone1": "99222222",
                "registrationNumber": "",
                "address": "",
            }
        ]

        response = self.client.post(
            "/api/state",
            {
                "state": device_b,
                "actor": {"id": "admin", "email": "admin@tomuda.mn"},
            },
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200, response.content)
        saved_ids = {c["id"] for c in response.json()["state"]["customers"]}
        self.assertIn("c-device-a", saved_ids)
        self.assertIn("c-device-b", saved_ids)

        row = AppState.objects.get(key="main")
        db_ids = {c["id"] for c in row.data["customers"]}
        self.assertIn("c-device-a", db_ids)
        self.assertIn("c-device-b", db_ids)

    def test_save_state_honors_customer_deletion_log(self):
        state = default_state()
        state["customers"] = [
            {
                "id": "c-keep",
                "name": "Keep",
                "companyName": "",
                "phone1": "99111111",
            },
            {
                "id": "c-delete",
                "name": "Delete me",
                "companyName": "",
                "phone1": "99333333",
            },
        ]
        AppState.objects.create(key="main", data=state)

        incoming = default_state()
        incoming["customers"] = [
            {
                "id": "c-keep",
                "name": "Keep",
                "companyName": "",
                "phone1": "99111111",
            }
        ]
        incoming["deletionLog"] = [
            {
                "type": "customer",
                "id": "c-delete",
                "deletedBy": "admin",
                "deletedAt": "2026-08-05T00:00:00",
            }
        ]

        response = self.client.post(
            "/api/state",
            {
                "state": incoming,
                "actor": {"id": "admin", "email": "admin@tomuda.mn"},
            },
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200, response.content)
        saved_ids = {c["id"] for c in response.json()["state"]["customers"]}
        self.assertIn("c-keep", saved_ids)
        self.assertNotIn("c-delete", saved_ids)

    def test_promotion_deletion_log_removes_rules_on_merge(self):
        from dashboard.state_merge import merge_app_states, promotion_rule_canonical_fingerprint

        rule = {
            "buyProductIds": ["p2", "p1"],
            "buyQtyByProduct": {"p2": 50, "p1": 50},
            "freeProductIds": ["p3"],
            "freeQty": 10,
            "buyMode": "each",
        }
        fp = promotion_rule_canonical_fingerprint(rule)
        remote = {
            "promotionRules": {"quantity": [rule], "price": [], "payment": []},
            "promotionDeletionLog": [],
        }
        local = {
            "promotionRules": {"quantity": [], "price": [], "payment": []},
            "promotionDeletionLog": [
                {
                    "kind": "quantity",
                    "fingerprint": fp,
                    "deletedBy": "admin",
                    "deletedAt": "2026-08-05T12:00:00",
                    "restored": False,
                    "updatedAt": "2026-08-05T12:00:00",
                }
            ],
        }
        merged = merge_app_states(remote, local)
        self.assertEqual(merged["promotionRules"]["quantity"], [])


@override_settings(SECURE_SSL_REDIRECT=False)
class CustomerUpsertApiTests(TestCase):
    def test_upsert_adds_customer_without_full_state_post(self):
        state = default_state()
        state["customers"] = [
            {
                "id": "c-existing",
                "name": "Existing",
                "companyName": "",
                "phone1": "99111111",
            }
        ]
        AppState.objects.create(key="main", data=state)

        response = self.client.post(
            "/api/customers/upsert",
            {
                "customer": {
                    "id": "c-new-device",
                    "name": "New from phone",
                    "companyName": "Store LLC",
                    "phone1": "99444444",
                    "registrationNumber": "",
                    "address": "UB",
                },
                "actor": {"id": "admin", "email": "admin@tomuda.mn"},
            },
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200, response.content)
        payload = response.json()
        self.assertEqual(payload["customer"]["id"], "c-new-device")
        self.assertEqual(payload["customer"]["name"], "New from phone")

        row = AppState.objects.get(key="main")
        db_ids = {c["id"] for c in row.data["customers"]}
        self.assertIn("c-new-device", db_ids)
        self.assertIn("c-existing", db_ids)

        # Peer device full-state save without the new customer must not wipe it.
        peer = default_state()
        peer["customers"] = [
            {
                "id": "c-existing",
                "name": "Existing",
                "companyName": "",
                "phone1": "99111111",
            }
        ]
        peer_response = self.client.post(
            "/api/state",
            {
                "state": peer,
                "actor": {"id": "admin", "email": "admin@tomuda.mn"},
            },
            content_type="application/json",
        )
        self.assertEqual(peer_response.status_code, 200, peer_response.content)
        peer_ids = {c["id"] for c in peer_response.json()["state"]["customers"]}
        self.assertIn("c-new-device", peer_ids)
        self.assertIn("c-existing", peer_ids)

    def test_upsert_name_and_map_survive_stale_peer_state_save(self):
        state = default_state()
        state["customers"] = [
            {
                "id": "c-ulz",
                "name": "Old store",
                "companyName": "",
                "phone1": "99111111",
                "latitude": "",
                "longitude": "",
            }
        ]
        AppState.objects.create(key="main", data=state)

        upsert = self.client.post(
            "/api/customers/upsert",
            {
                "customer": {
                    "id": "c-ulz",
                    "name": "Улз",
                    "companyName": "",
                    "phone1": "99111111",
                    "latitude": 47.918,
                    "longitude": 106.917,
                },
                "actor": {"id": "admin", "email": "admin@tomuda.mn"},
            },
            content_type="application/json",
        )
        self.assertEqual(upsert.status_code, 200, upsert.content)
        saved = upsert.json()["customer"]
        self.assertEqual(saved["id"], "c-ulz")
        self.assertEqual(saved["name"], "Улз")
        self.assertEqual(float(saved["latitude"]), 47.918)
        self.assertEqual(float(saved["longitude"]), 106.917)
        self.assertTrue(saved.get("updatedAt"))

        row = AppState.objects.get(key="main")
        db_customer = next(c for c in row.data["customers"] if c["id"] == "c-ulz")
        self.assertEqual(db_customer["name"], "Улз")
        self.assertEqual(float(db_customer["latitude"]), 47.918)
        self.assertEqual(float(db_customer["longitude"]), 106.917)

        peer = default_state()
        peer["customers"] = [
            {
                "id": "c-ulz",
                "name": "Old store",
                "companyName": "",
                "phone1": "99111111",
                "latitude": "",
                "longitude": "",
            }
        ]
        peer_response = self.client.post(
            "/api/state",
            {
                "state": peer,
                "actor": {"id": "admin", "email": "admin@tomuda.mn"},
            },
            content_type="application/json",
        )
        self.assertEqual(peer_response.status_code, 200, peer_response.content)
        after = next(
            c
            for c in peer_response.json()["state"]["customers"]
            if c["id"] == "c-ulz"
        )
        self.assertEqual(after["name"], "Улз")
        self.assertEqual(float(after["latitude"]), 47.918)
        self.assertEqual(float(after["longitude"]), 106.917)


@override_settings(SECURE_SSL_REDIRECT=False)
class ProductUpsertApiTests(TestCase):
    def test_upsert_adds_product_and_keeps_stock_on_edit(self):
        state = default_state()
        state["products"] = [
            {
                "id": "p-existing",
                "name": "Existing",
                "category": "Ундаа",
                "unit": "ширхэг",
                "boxQuantity": 24,
                "price": 1000,
                "stock": 73,
                "costPrice": 500,
                "minStock": 0,
            }
        ]
        AppState.objects.create(key="main", data=state)

        create_response = self.client.post(
            "/api/products/upsert",
            {
                "product": {
                    "id": "p-new-device",
                    "name": "New snack",
                    "category": "Амттан",
                    "unit": "ширхэг",
                    "boxQuantity": 24,
                    "largeBoxQuantity": 288,
                    "price": 2450,
                    "country": "Germany",
                },
                "actor": {"id": "admin", "email": "admin@tomuda.mn"},
            },
            content_type="application/json",
        )
        self.assertEqual(create_response.status_code, 200, create_response.content)
        created = create_response.json()["product"]
        self.assertEqual(created["id"], "p-new-device")
        self.assertEqual(created["largeBoxQuantity"], 288)
        self.assertEqual(created.get("stock", 0), 0)

        edit_response = self.client.post(
            "/api/products/upsert",
            {
                "product": {
                    "id": "p-existing",
                    "name": "Existing renamed",
                    "category": "Ундаа",
                    "unit": "ширхэг",
                    "boxQuantity": 24,
                    "largeBoxQuantity": 12,
                    "price": 1100,
                    "country": "Монгол",
                },
                "actor": {"id": "admin", "email": "admin@tomuda.mn"},
            },
            content_type="application/json",
        )
        self.assertEqual(edit_response.status_code, 200, edit_response.content)
        edited = edit_response.json()["product"]
        self.assertEqual(edited["name"], "Existing renamed")
        self.assertEqual(edited["largeBoxQuantity"], 12)
        self.assertEqual(edited["stock"], 73)
        self.assertEqual(edited["costPrice"], 500)

        row = AppState.objects.get(key="main")
        by_id = {p["id"]: p for p in row.data["products"]}
        self.assertIn("p-new-device", by_id)
        self.assertEqual(by_id["p-existing"]["stock"], 73)


@override_settings(SECURE_SSL_REDIRECT=False)
class OrderUpsertApiTests(TestCase):
    def test_upsert_adds_order_and_survives_peer_state_save(self):
        state = default_state()
        state["products"] = [
            {
                "id": "p1",
                "name": "Cola",
                "price": 1000,
                "stock": 50,
                "unit": "ширхэг",
            }
        ]
        state["orders"] = []
        AppState.objects.create(key="main", data=state)

        order = {
            "id": "o-device-b",
            "customerId": "c1",
            "customerName": "Шинэ хүнс 2",
            "items": [
                {
                    "productId": "p1",
                    "productName": "Cola",
                    "quantity": 5,
                    "price": 1000,
                    "total": 5000,
                }
            ],
            "total": 5000,
            "grossTotal": 5000,
            "discountAmount": 0,
            "status": "pending",
            "paymentTerm": "credit",
            "isPaid": False,
            "employeeId": "admin",
            "employeeName": "Admin",
            "createdAt": "2026-08-05T07:00:00.000Z",
            "deliveryDate": "2026-08-05",
        }
        response = self.client.post(
            "/api/orders/upsert",
            {
                "order": order,
                "actor": {"id": "admin", "email": "admin@tomuda.mn"},
            },
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200, response.content)
        payload = response.json()
        self.assertEqual(payload["order"]["id"], "o-device-b")
        self.assertEqual(
            next(p["stock"] for p in payload["state"]["products"] if p["id"] == "p1"),
            45,
        )

        # Peer device posts an older empty order list — merge must keep the order.
        peer = default_state()
        peer["products"] = [
            {
                "id": "p1",
                "name": "Cola",
                "price": 1000,
                "stock": 50,
                "unit": "ширхэг",
            }
        ]
        peer["orders"] = []
        peer_response = self.client.post(
            "/api/state",
            {
                "state": peer,
                "actor": {"id": "admin", "email": "admin@tomuda.mn"},
            },
            content_type="application/json",
        )
        self.assertEqual(peer_response.status_code, 200, peer_response.content)
        peer_ids = {o["id"] for o in peer_response.json()["state"]["orders"]}
        self.assertIn("o-device-b", peer_ids)

        row = AppState.objects.get(key="main")
        self.assertIn("o-device-b", {o["id"] for o in row.data["orders"]})

