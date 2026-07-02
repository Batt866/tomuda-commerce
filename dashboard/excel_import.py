"""Excel import for customers and products (AppState JSON)."""

from __future__ import annotations

import io
import json
import re
import time
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CUSTOMER_HEADERS = [
    "Нэр",
    "Регистр",
    "Утас 1",
    "Утас 2",
    "Аймаг / Хот",
    "Дүүрэг",
    "Хороо",
    "Дэлгэрэнгүй хаяг",
]

PRODUCT_HEADERS = [
    "Barcode",
    "Барааны нэр",
    "Хэмжих нэгж",
    "Борлуулалтын үнэ",
    "Үйлдвэрлэсэн үнэ",
    "Үйлдвэрлэсэн улс",
    "Зураг URL",
]

CUSTOMER_HEADER_ALIASES: dict[str, str] = {
    "нэр": "name",
    "name": "name",
    "байгууллагын нэр": "name",
    "company": "name",
    "company name": "name",
    "компани": "name",
    "регистр": "registrationNumber",
    "registration": "registrationNumber",
    "рд": "registrationNumber",
    "rd": "registrationNumber",
    "рдад": "registrationNumber",
    "утас 1": "phone1",
    "утас1": "phone1",
    "утас": "phone1",
    "гар утас": "phone1",
    "phone1": "phone1",
    "phone": "phone1",
    "mobile": "phone1",
    "утас 2": "phone2",
    "утас2": "phone2",
    "phone2": "phone2",
    "аймаг / хот": "province",
    "аймаг/хот": "province",
    "аймаг": "province",
    "хот": "province",
    "province": "province",
    "дүүрэг": "district",
    "дүүрэг/сум": "district",
    "сум": "district",
    "district": "district",
    "хороо": "khoroo",
    "khoroo": "khoroo",
    "дэлгэрэнгүй хаяг": "address",
    "хаяг": "address",
    "address": "address",
}

PRODUCT_HEADER_ALIASES: dict[str, str] = {
    "barcode": "barcode",
    "баркод": "barcode",
    "барааны нэр": "name",
    "нэр": "name",
    "name": "name",
    "хэмжих нэгж": "unit",
    "нэгж": "unit",
    "unit": "unit",
    "борлуулалтын үнэ": "price",
    "price": "price",
    "үнэ": "price",
    "үртөг үнэ": "costPrice",
    "өртөг үнэ": "costPrice",
    "costprice": "costPrice",
    "cost price": "costPrice",
    "үйлдвэрлэсэн үнэ": "costPrice",
    "үйлдвэрлэсэн улс": "country",
    "улс": "country",
    "country": "country",
    "зураг": "image",
    "зураг url": "image",
    "зургийн холбоос": "image",
    "image": "image",
    "image url": "image",
    "image_url": "image",
    "photo": "image",
    "picture": "image",
}

UNIT_ALIASES: dict[str, str] = {
    "кг": "KG",
    "kg": "KG",
    "KG": "KG",
    "ширхэг": "ширхэг",
    "ш": "ширхэг",
    "ш.": "ширхэг",
    "метр": "метр",
    "m": "метр",
    "м": "метр",
}

EXCEL_IMPORT_FORMAT_HINT = (
    "Файлын формат .xlsx байх ёстой. «Формат татах» товчоор татсан .xlsx "
    "загвар ашиглана уу. Хуучин .xls файл мөн дэмжигдэнэ."
)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", _cell_text(value).lower())


def _image_url_text(value: Any) -> str:
    text = _cell_text(value)
    return "" if text.lower() in {"", "nan", "none", "null"} else text


def _registration_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def _valid_phone(value: str) -> bool:
    if not value:
        return True
    digits = re.sub(r"\D", "", value)
    return len(digits) >= 6


def _parse_number(value: Any, *, allow_empty: bool = False) -> tuple[float | None, str | None]:
    text = _cell_text(value)
    if not text:
        return (None, None) if allow_empty else (None, "тоон утга оруулна уу")
    cleaned = text.replace(" ", "").replace(",", "")
    try:
        number = float(cleaned)
    except ValueError:
        return None, "тоон утга биш байна"
    if number < 0:
        return None, "сөрөг утга зөвшөөрөгдөхгүй"
    return number, None


def _normalize_unit(value: str) -> tuple[str | None, str | None]:
    text = _cell_text(value)
    if not text:
        return "ширхэг", None
    key = text.lower().replace(".", "")
    normalized = UNIT_ALIASES.get(key) or UNIT_ALIASES.get(text) or UNIT_ALIASES.get(text.lower())
    if normalized:
        return normalized, None
    if text in ("ширхэг", "KG", "метр"):
        return text, None
    return None, f"Хэмжих нэгж буруу байна ({text})"


def _row_is_empty(cells: list[str]) -> bool:
    return not any(cells)


def _trim_empty_tail(rows: list[list[str]]) -> list[list[str]]:
    while rows and _row_is_empty(rows[-1]):
        rows.pop()
    return rows


def _is_template_example_customer(name: str, reg_raw: str) -> bool:
    normalized = _normalize_header(name)
    if normalized in {"жишээ ххк", "жишээ", "example", "sample"}:
        return True
    reg_digits = _registration_digits(reg_raw)
    return normalized == "жишээ ххк" or reg_digits == "1234567"


def _load_rows_xlsx(data: bytes) -> list[list[str]]:
    wb = load_workbook(io.BytesIO(data), read_only=False, data_only=True)
    try:
        rows: list[list[str]] = []
        for row in wb.active.iter_rows(
            min_row=1,
            max_row=wb.active.max_row,
            max_col=wb.active.max_column,
            values_only=True,
        ):
            rows.append([_cell_text(v) for v in row])
        return _trim_empty_tail(rows)
    finally:
        wb.close()


def _openpyxl_sheet_rows(ws: Any) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        max_col=ws.max_column,
        values_only=True,
    ):
        rows.append([_cell_text(v) for v in row])
    return _trim_empty_tail(rows)


def _sheet_import_score(rows: list[list[str]]) -> int:
    if not rows:
        return -1

    score = 0
    header_idx, _ = _find_header_row(rows, CUSTOMER_HEADER_ALIASES, ("name",))
    product_header_idx, _ = _find_header_row(
        rows,
        PRODUCT_HEADER_ALIASES,
        ("barcode", "name"),
    )
    if header_idx >= 0:
        score += 10
        score += sum(1 for row in rows[header_idx + 1 :] if not _row_is_empty(row))
    if product_header_idx >= 0:
        score += 10
        score += sum(
            1 for row in rows[product_header_idx + 1 :] if not _row_is_empty(row)
        )
    if score:
        return score
    return 1 if any(not _row_is_empty(row) for row in rows) else -1


def _best_rows_from_sheets(sheet_rows: list[list[list[str]]]) -> list[list[str]]:
    best_rows: list[list[str]] = []
    best_score = -1
    for rows in sheet_rows:
        score = _sheet_import_score(rows)
        if score > best_score:
            best_score = score
            best_rows = rows
    return best_rows


def _best_rows_xlsx(data: bytes) -> list[list[str]]:
    wb = load_workbook(io.BytesIO(data), read_only=False, data_only=True)
    try:
        best_rows = _best_rows_from_sheets(
            [_openpyxl_sheet_rows(ws) for ws in wb.worksheets]
        )
        return best_rows or _openpyxl_sheet_rows(wb.active)
    finally:
        wb.close()


def _load_workbook_rows(data: bytes, filename: str) -> list[list[str]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _best_rows_xlsx(data)
    if name.endswith(".xls"):
        return _load_rows_xls(data)
    raise ValueError(EXCEL_IMPORT_FORMAT_HINT)


def _best_sheet_rows(data: bytes, filename: str) -> list[list[str]]:
    name = (filename or "").lower()
    if not name.endswith(".xlsx"):
        return _load_workbook_rows(data, filename)

    return _best_rows_xlsx(data)


def _load_rows_xls(data: bytes) -> list[list[str]]:
    if data.startswith(b"PK"):
        return _best_rows_xlsx(data)
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError(
            f"XLS унших xlrd суулгаагүй байна. {EXCEL_IMPORT_FORMAT_HINT}"
        ) from exc
    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception as exc:
        raise ValueError(
            f"XLS файл уншиж чадсангүй. {EXCEL_IMPORT_FORMAT_HINT}"
        ) from exc

    sheets: list[list[list[str]]] = []
    for sheet_idx in range(book.nsheets):
        sheet = book.sheet_by_index(sheet_idx)
        rows: list[list[str]] = []
        for r in range(sheet.nrows):
            rows.append([_cell_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)])
        sheets.append(_trim_empty_tail(rows))
    return _best_rows_from_sheets(sheets)


def load_sheet_rows(data: bytes, filename: str) -> list[list[str]]:
    return _best_sheet_rows(data, filename)


def _map_headers(header_row: list[str], aliases: dict[str, str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = aliases.get(_normalize_header(cell))
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def _find_header_row(
    rows: list[list[str]],
    aliases: dict[str, str],
    required: tuple[str, ...],
    *,
    scan_limit: int = 20,
) -> tuple[int, dict[str, int]]:
    for idx, row in enumerate(rows[:scan_limit]):
        mapping = _map_headers(row, aliases)
        if all(key in mapping for key in required):
            return idx, mapping
    return -1, {}


def _row_dict(row: list[str], mapping: dict[str, int]) -> dict[str, str]:
    out: dict[str, str] = {}
    for field, idx in mapping.items():
        out[field] = row[idx] if idx < len(row) else ""
    return out


def _excel_col_width(*values: Any, minimum: float = 12) -> float:
    best = minimum
    for value in values:
        text = _cell_text(value)
        if not text:
            continue
        units = 0.0
        for ch in text:
            units += 1.5 if ord(ch) > 127 else 1.0
        best = max(best, units + 5)
    return min(best, 80)


def _format_import_worksheet(
    ws,
    headers: list[str],
    example_row: list[Any],
    *,
    text_columns: set[int] | None = None,
    min_col_widths: list[float] | None = None,
) -> None:
    text_columns = text_columns or set()
    mins = min_col_widths or [12] * len(headers)
    col_widths = [
        _excel_col_width(
            headers[i],
            example_row[i] if i < len(example_row) else "",
            minimum=mins[i] if i < len(mins) else 12,
        )
        for i in range(len(headers))
    ]
    header_font = Font(bold=True, size=11, color="000000")
    header_fill = PatternFill("solid", fgColor="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="center", wrap_text=True, horizontal="left")

    longest_header_lines = 1
    for col_idx, width in enumerate(col_widths, start=1):
        header_text = headers[col_idx - 1]
        units = sum(1.4 if ord(ch) > 127 else 1.0 for ch in header_text)
        lines = max(1, int(units // max(width - 2, 8)) + (1 if units > width else 0))
        longest_header_lines = max(longest_header_lines, lines)

    ws.row_dimensions[1].height = max(36, 18 * longest_header_lines + 8)
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A2"

    for col_idx, width in enumerate(col_widths, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width
        header_cell = ws.cell(row=1, column=col_idx, value=headers[col_idx - 1])
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_align

        value = example_row[col_idx - 1] if col_idx - 1 < len(example_row) else ""
        body_cell = ws.cell(row=2, column=col_idx, value=value)
        body_cell.alignment = body_align
        if col_idx in text_columns:
            body_cell.number_format = "@"


def build_customer_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Харилцагч"
    example = [
        "Жишээ ХХК",
        "1234567",
        "99112233",
        "",
        "Улаанбаатар",
        "Баянзүрх",
        "1",
        "1-р хороо, 10-р байр",
    ]
    _format_import_worksheet(
        ws,
        CUSTOMER_HEADERS,
        example,
        text_columns={2, 3, 4},
        min_col_widths=[18, 14, 14, 14, 20, 16, 12, 34],
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_product_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Бараа"
    example = ["6977236071316", "Жишээ бараа", "ширхэг", 15000, 10000, "Монгол", ""]
    _format_import_worksheet(
        ws,
        PRODUCT_HEADERS,
        example,
        text_columns={1, 7},
        min_col_widths=[18, 26, 20, 30, 30, 22, 42],
    )
    for col_idx in (4, 5):
        ws.cell(row=2, column=col_idx).number_format = "#,##0"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _append_import_log(
    state: dict[str, Any],
    *,
    kind: str,
    actor: dict[str, Any] | None,
    report: dict[str, Any],
) -> None:
    logs = list(state.get("importLogs") or [])
    logs.insert(
        0,
        {
            "id": f"import-{int(time.time() * 1000)}",
            "type": kind,
            "createdAt": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "actorId": str((actor or {}).get("id") or ""),
            "actorEmail": str((actor or {}).get("email") or ""),
            **report,
        },
    )
    state["importLogs"] = logs[:50]


def import_customers_into_state(
    state: dict[str, Any],
    rows: list[list[str]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    if not rows:
        raise ValueError(f"Excel хоосон байна. {EXCEL_IMPORT_FORMAT_HINT}")
    header_idx, header_map = _find_header_row(rows, CUSTOMER_HEADER_ALIASES, ("name",))
    if header_idx < 0:
        raise ValueError(
            "Excel-д «Нэр» багана олдсонгүй. "
            f"{EXCEL_IMPORT_FORMAT_HINT}"
        )
    data_rows = rows[header_idx + 1 :]

    existing_by_reg = {
        _registration_digits(str(c.get("registrationNumber") or "")): c
        for c in state.get("customers") or []
        if _registration_digits(str(c.get("registrationNumber") or ""))
    }
    batch_regs: set[str] = set()
    customers = list(state.get("customers") or [])
    errors: list[dict[str, Any]] = []
    success = 0
    updated = 0
    created = 0
    total = 0
    now = int(time.time() * 1000)

    for offset, row in enumerate(data_rows, start=header_idx + 2):
        if _row_is_empty(row):
            continue
        data = _row_dict(row, header_map)
        name = _cell_text(data.get("name"))
        reg_raw = _cell_text(data.get("registrationNumber"))
        reg_digits = _registration_digits(reg_raw)
        phone1 = _cell_text(data.get("phone1"))
        phone2 = _cell_text(data.get("phone2"))

        if _is_template_example_customer(name, reg_raw):
            continue
        total += 1
        if not name:
            errors.append({"row": offset, "message": "Нэр хоосон байна."})
            continue
        if reg_digits and reg_digits in batch_regs:
            errors.append({"row": offset, "message": "Регистр давхардаж байна."})
            continue
        if phone1 and not _valid_phone(phone1):
            errors.append({"row": offset, "message": "Утасны формат буруу байна."})
            continue
        if phone2 and not _valid_phone(phone2):
            errors.append({"row": offset, "message": "Утас 2 формат буруу байна."})
            continue

        existing = existing_by_reg.get(reg_digits) if reg_digits else None
        if existing:
            # Excel-ээр давтан оруулахад давхар бичлэг үүсгэхгүй, байгаа харилцагчийг шинэчилнэ.
            existing["name"] = name
            existing["companyName"] = name
            existing["registrationNumber"] = reg_raw
            existing["phone1"] = phone1
            existing["phone2"] = phone2
            existing["province"] = _cell_text(data.get("province")) or "Улаанбаатар"
            existing["district"] = _cell_text(data.get("district"))
            existing["khoroo"] = _cell_text(data.get("khoroo"))
            existing["address"] = _cell_text(data.get("address"))
            updated += 1
        else:
            now += 1
            customer = {
                "id": str(now),
                "name": name,
                "registrationNumber": reg_raw,
                "companyName": name,
                "phone1": phone1,
                "phone2": phone2,
                "province": _cell_text(data.get("province")) or "Улаанбаатар",
                "district": _cell_text(data.get("district")),
                "khoroo": _cell_text(data.get("khoroo")),
                "address": _cell_text(data.get("address")),
                "latitude": "",
                "longitude": "",
                "locationText": "",
                "image": "",
            }
            customers.append(customer)
            created += 1
            if reg_digits:
                existing_by_reg[reg_digits] = customer
        if reg_digits:
            batch_regs.add(reg_digits)
        success += 1

    if total <= 0:
        raise ValueError(
            "Excel-д өгөгдөлтэй мөр олдсонгүй. "
            f"{EXCEL_IMPORT_FORMAT_HINT}"
        )

    next_state = dict(state)
    next_state["customers"] = customers
    report = {
        "total": total,
        "success": success,
        "created": created,
        "updated": updated,
        "failed": len(errors),
        "errors": errors,
    }
    return next_state, report


def import_products_into_state(
    state: dict[str, Any],
    rows: list[list[str]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    if not rows:
        raise ValueError(f"Excel хоосон байна. {EXCEL_IMPORT_FORMAT_HINT}")
    header_idx, header_map = _find_header_row(
        rows,
        PRODUCT_HEADER_ALIASES,
        ("barcode", "name"),
    )
    if header_idx < 0:
        raise ValueError(
            "Excel-д «Barcode» болон «Барааны нэр» багана шаардлагатай. "
            f"{EXCEL_IMPORT_FORMAT_HINT}"
        )
    data_rows = rows[header_idx + 1 :]

    batch_barcodes: set[str] = set()
    products = [
        dict(p) if isinstance(p, dict) else p for p in state.get("products") or []
    ]
    existing_by_barcode = {
        re.sub(r"\D", "", str(p.get("barcode") or "")): p
        for p in products
        if isinstance(p, dict) and re.sub(r"\D", "", str(p.get("barcode") or ""))
    }
    errors: list[dict[str, Any]] = []
    touched_product_ids: list[str] = []
    success = 0
    updated = 0
    created = 0
    total = 0
    now = int(time.time() * 1000)

    for offset, row in enumerate(data_rows, start=header_idx + 2):
        if _row_is_empty(row):
            continue
        total += 1
        data = _row_dict(row, header_map)
        barcode = re.sub(r"\D", "", _cell_text(data.get("barcode")))
        name = _cell_text(data.get("name"))
        unit_raw = _cell_text(data.get("unit"))
        price, price_err = _parse_number(data.get("price"), allow_empty=False)
        cost, cost_err = _parse_number(data.get("costPrice"), allow_empty=True)

        if not barcode:
            errors.append({"row": offset, "message": "Barcode хоосон байна."})
            continue
        if not name:
            errors.append({"row": offset, "message": "Барааны нэр хоосон байна."})
            continue
        if barcode in batch_barcodes:
            errors.append({"row": offset, "message": "Barcode давхардаж байна."})
            continue
        if price_err:
            errors.append({"row": offset, "message": "Үнэ тоон утга биш байна."})
            continue
        if cost_err:
            errors.append({"row": offset, "message": "Үйлдвэрлэсэн үнэ тоон утга биш байна."})
            continue
        unit, unit_err = _normalize_unit(unit_raw)
        if unit_err:
            errors.append({"row": offset, "message": "Хэмжих нэгж буруу байна."})
            continue

        if barcode in existing_by_barcode:
            errors.append(
                {
                    "row": offset,
                    "message": "Али хэдийн бүртгэгдсэн product байна.",
                }
            )
            continue

        country = _cell_text(data.get("country")) or "Монгол"
        image_source = _image_url_text(data.get("image"))

        now += 1
        pid = str(now)
        product = {
            "id": pid,
            "barcode": barcode,
            "name": name,
            "category": "Бусад",
            "unit": unit or "ширхэг",
            "boxQuantity": 1,
            "price": int(price or 0),
            "costPrice": int(cost or 0),
            "stock": 0,
            "minStock": 0,
            "country": country,
            "image": image_source,
        }
        products.append(product)
        existing_by_barcode[barcode] = product
        created += 1
        touched_product_ids.append(pid)
        batch_barcodes.add(barcode)
        success += 1

    next_state = dict(state)
    next_state["products"] = products
    report = {
        "total": total,
        "success": success,
        "created": created,
        "updated": updated,
        "failed": len(errors),
        "errors": errors,
        "_productIds": touched_product_ids,
    }
    return next_state, report


def parse_actor(raw: str | None) -> dict[str, Any] | None:
    if not raw:
        return None
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return None
    return data if isinstance(data, dict) else None
